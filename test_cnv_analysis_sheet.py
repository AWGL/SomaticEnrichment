import unittest

from cnv_analysis_sheet import *
import openpyxl

path="./tests/"


wb=Workbook()
ws1=wb.create_sheet("Patient demographics")
ws2=wb.create_sheet("NTC and Reads check")
ws3=wb.create_sheet("hotspot_cnvs")
ws4=wb.create_sheet("Lists")

class test_make_worksheets(unittest.TestCase):

	def test_fish_probe(self):
		self.assertEqual(fish_probe("8-385639-596739"),None)
		self.assertEqual(fish_probe("1-385639-596739"),"NO")
		self.assertEqual(fish_probe("1-3397084-3766385"),"YES")
		self.assertEqual(fish_probe("1-3397089-3766379"),"YES")
		self.assertEqual(fish_probe("19-47924004-47924009"),"YES")


	def test_get_CNV_file(self):

		cnvs, ws1_output, ws2_output, ws3_output, ws4_output=(get_CNV_file(path, "sample1",ws1,ws2,ws3,ws4))

		self.assertEqual(ws3_output["A10"].value,"gene/region")
		self.assertEqual(ws3_output["B10"].value,"chromosome")
		self.assertEqual(ws3_output["C10"].value,"start")
		self.assertEqual(ws3_output["D10"].value,"end")
		self.assertEqual(ws3_output["E10"].value,"log2")
		self.assertEqual(ws3_output["F10"].value,"depth")
		self.assertEqual(ws3_output["G10"].value,"weight")
		self.assertEqual(ws3_output["H10"].value,"baf")
		self.assertEqual(ws3_output["I10"].value,"ci_hi")
		self.assertEqual(ws3_output["J10"].value,"ci_lo")
		self.assertEqual(ws3_output["K10"].value,"segment_probes")
		self.assertEqual(ws3_output["L10"].value,"Fish probe region?")


		self.assertEqual(ws3_output["A11"].value,"1p")
		self.assertEqual(ws3_output["B11"].value,"1")
		self.assertEqual(ws3_output["C11"].value,"5974")
		self.assertEqual(ws3_output["D11"].value,"586947")
		self.assertEqual(ws3_output["E11"].value,-0.4)
		self.assertEqual(ws3_output["F11"].value, 438)
		self.assertEqual(ws3_output["G11"].value, 27.1)
		self.assertEqual(ws3_output["H11"].value, 0.5385)
		self.assertEqual(ws3_output["I11"].value, 76.23)
		self.assertEqual(ws3_output["J11"].value, 65.34)
		self.assertEqual(ws3_output["K11"].value,22)
		self.assertEqual(ws3_output["L11"].value,"NO")


		self.assertEqual(ws3_output["A12"].value,"1p")
		self.assertEqual(ws3_output["B12"].value,"1")
		self.assertEqual(ws3_output["C12"].value,"3397084")
		self.assertEqual(ws3_output["D12"].value,"3766385")
		self.assertEqual(ws3_output["E12"].value,-0.43)
		self.assertEqual(ws3_output["F12"].value, 436)
		self.assertEqual(ws3_output["G12"].value, 27.5)
		self.assertEqual(ws3_output["H12"].value, 0.5647)
		self.assertEqual(ws3_output["I12"].value, 76.85)
		self.assertEqual(ws3_output["J12"].value, 61.33)
		self.assertEqual(ws3_output["K12"].value,21)
		self.assertEqual(ws3_output["L12"].value,"YES")


		self.assertEqual(ws3_output["A13"].value,"19q")
		self.assertEqual(ws3_output["B13"].value,"19")
		self.assertEqual(ws3_output["C13"].value,"48023008")
		self.assertEqual(ws3_output["D13"].value,"48023888")
		self.assertEqual(ws3_output["E13"].value,-0.32)
		self.assertEqual(ws3_output["F13"].value, 534)
		self.assertEqual(ws3_output["G13"].value, 33.1)
		self.assertEqual(ws3_output["H13"].value, 0.5539999999999999)
		self.assertEqual(ws3_output["I13"].value, 7.85)
		self.assertEqual(ws3_output["J13"].value, 12.63)
		self.assertEqual(ws3_output["K13"].value,29)
		self.assertEqual(ws3_output["L13"].value,"YES")


		self.assertEqual(ws3_output["A14"].value,"19q")
		self.assertEqual(ws3_output["B14"].value,"19")
		self.assertEqual(ws3_output["C14"].value,"48024003")
		self.assertEqual(ws3_output["D14"].value,"48374582")
		self.assertEqual(ws3_output["E14"].value,-0.78)
		self.assertEqual(ws3_output["F14"].value, 111)
		self.assertEqual(ws3_output["G14"].value, 12.3)
		self.assertEqual(ws3_output["H14"].value, 0.867)
		self.assertEqual(ws3_output["I14"].value, 23.85)
		self.assertEqual(ws3_output["J14"].value, 81.33)
		self.assertEqual(ws3_output["K14"].value,14)
		self.assertEqual(ws3_output["L14"].value,"YES")

		self.assertEqual(ws3_output["A24"].value,"Analysis")



	def test_get_coverage(self):

		coverage, ws1_output, ws2_output, ws3_output, ws4_output=(get_coverage(path, "seqid","sample1",ws1,ws2,ws3,ws4))

		self.assertEqual(ws2_output["A1"].value,"Locus")
		self.assertEqual(ws2_output["B1"].value,"Depth_for_sample1")
		self.assertEqual(ws2_output["C1"].value,"Depth_for_NTC")
		self.assertEqual(ws2_output["D1"].value,"%NTC")

		self.assertEqual(ws2_output["A2"].value,"1:836814")
		self.assertEqual(ws2_output["B2"].value,5.0)
		self.assertEqual(ws2_output["C2"].value,0.0)
		self.assertEqual(ws2_output["D2"].value,0.0)

		self.assertEqual(ws2_output["A3"].value,"1:836815")
		self.assertEqual(ws2_output["B3"].value,10.0)
		self.assertEqual(ws2_output["C3"].value,1.0)
		self.assertEqual(ws2_output["D3"].value,10)

		self.assertEqual(ws2_output["H13"].value,2000001)


		self.assertEqual(ws2_output["A35465"].value,"19:58729903")
		self.assertEqual(ws2_output["B35465"].value,100.0)
		self.assertEqual(ws2_output["C35465"].value,0.0)
		self.assertEqual(ws2_output["D35465"].value,0.0)

		self.assertEqual(ws2_output["A35466"].value,"19:58729904")
		self.assertEqual(ws2_output["B35466"].value,20.0)
		self.assertEqual(ws2_output["C35466"].value,1.0)
		self.assertEqual(ws2_output["D35466"].value, 5)

		self.assertEqual(ws2_output["A35467"].value,"19:58729905")
		self.assertEqual(ws2_output["B35467"].value,50.0)
		self.assertEqual(ws2_output["C35467"].value,2.0)
		self.assertEqual(ws2_output["D35467"].value, 4)

		self.assertEqual(ws2_output["A35468"].value,None)
		self.assertEqual(ws2_output["B35468"].value,None)
		self.assertEqual(ws2_output["C35468"].value,None)
		self.assertEqual(ws2_output["D35468"].value,None)





















