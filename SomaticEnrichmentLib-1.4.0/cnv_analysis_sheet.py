import pandas
from openpyxl import Workbook
import os
import numpy
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM, BORDER_THIN, BORDER_THICK
from openpyxl.styles import Font
import argparse
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment
import datetime

wb=Workbook()
ws1=wb.create_sheet("Patient demographics")
ws2=wb.create_sheet("NTC and Reads check")
ws3=wb.create_sheet("hotspot_cnvs")
ws4=wb.create_sheet("Lists")


def populate_cells(seqId, sampleid, worksheetid, ws1, ws2, ws3, ws4):

    '''
    This function adds labels to the analysis sheets 
    '''

    ws1['A1']= "Date Received"
    ws1['B1']= "Date Requested"
    ws1['C1']= "Due Date"
    ws1['D1']= "LABNO"
    ws1['E1']= "Patient name"
    ws1['F1']= "DOB"
    ws1['G1']= "Reason for referral"
    ws1['H1']= "Referring Clinician"
    ws1['I1']= "Hospital"
    ws1['J1']= "Date reported"
    ws1['K1']= "TAT"
    ws1['L1']= "No of days in Histo"
    ws1['M1']= "Block/Slide/DNA"
    ws1['N1']= "% Tumour"
    ws1['O1']= "Result"
    ws1['P1']= "NGS worksheet"
    ws1['Q1']= "Qubit DNA conc. (ng/ul)"
    ws1['R1']= "Total DNA input"
    ws1['S1']= "Post PCR1 Qubit"
    ws1['T1']= "Date of NextSeq run"
    ws1['U1']= "NextSeq run ID"
    ws1['V1']= "Comments"

    ws1['D2']= sampleid
    ws1['P2']= worksheetid
    ws1['U2']= seqId

    dv = DataValidation(type="list", formula1='"YES, NO"', allow_blank=True)
    ws1.add_data_validation(dv)
    dv.add('B6')
    dv.add('C6')

    ws1['A7']= "Demographics and 1p19q request checked by:"
    ws1['A8']= "Checker 1 initials and date"
    ws1['A9']= "Checker 2 initials and date"

    ws1['A12']= "% Gliomas on run"
    ws1['A13']= "No. Samples on Pan Cancer run"
    ws1['A14']= "No. Glioma samples on Pan Cancer run"
    ws1['A15']= "% Glioma's on Pan Cancer run"
    ws1['A16']= "Checker 1 initials and date"
    ws1['A17']= "Checker 2 initials and date"
    ws1['B15']='=(B14/B13)*100'

    ws3['A3']= "Patient information"
    ws3['A4']= "Lab number"
    ws3['B4']= "Patient Name"
    ws3['C4']= "Tumour %"
    ws3['D4']= "Analysis"
    ws3['E4']= "Due date"

    ws3['A7']= "NGS wks"
    ws3['B7']= "NextSeq runId"
    ws3['C7']= "Total Reads check 1"
    ws3['D7']= "Total Reads check 2"
    ws3['E7']= "NTC check 1"
    ws3['F7']= "NTC check 2"

    ws3['A5']= sampleid
    ws3['B5']= "='Patient demographics'!E2"
    ws3['C5']= "='Patient demographics'!N2"
    ws3['D5']= "='Patient demographics'!G2"
    ws3['E5']= "='Patient demographics'!C2"
    ws3['E5'].number_format='mm-dd-yy'

    ws3['A8']= worksheetid
    ws3['B8']= seqId
    ws3['C8']= "='NTC and Reads check'!H16"
    ws3['D8']= "='NTC and Reads check'!H17"
    ws3['E8']= "='NTC and Reads check'!H4"
    ws3['F8']= "='NTC and Reads check'!H5"

    ws4['A1']= "No evidence of loss of the 1p36 or 19q13 loci"
    ws4['A2']= "Co-deletion 1p/19q"
    ws4['A3']= "Deletion of 1p without loss of 19q"
    ws4['A4']= "Deletion of 19q without loss of 1p"
    ws4['A5']= "Equivocal NGS result. FISH recommended"
    ws4['A6']= "Failed NGS. FISH recommended"
    ws4['A7']= "Complex pattern of loss/gain"
    ws4['A8']= "<50%T and no loss/threshold not met for loss of 1p and/or 19q"
    ws4['A9']= "<2000000 reads"


    ws4['B1']= "No loss of FISH probe on region 1p"
    ws4['B2']= "No loss of FISH probe on region 19q"
    ws4['B3']= "Loss below threshold FISH probe region 1p36"
    ws4['B4']= "Loss below threshold FISH probe region 19q13"
    ws4['B5']= "FISH probe region deleted on 1p"
    ws4['B6']= "FISH probe region deleted on 19q"
    ws4['B7']= "Failed NGS. FISH recommended"

    ws4.column_dimensions['A'].width=30
    ws4.column_dimensions['B'].width=30

    return ws1, ws2, ws3, ws4


def fish_probe(chr_start_end):

    '''
    Called by get_CNV_file function to highlight the cnvs within the fish probe region
    '''

    chr= (chr_start_end.split("-"))[0]
    start= (chr_start_end.split("-"))[1]
    end= (chr_start_end.split("-"))[2]

    #fish_probe_regions
    start_1p=3397084
    end_1p=3766783
    start_19q=47924005
    end_19q=48374711
    start=float(start)
    end=float(end)


    if (chr=="1"):
        if ((start<=start_1p)and (end>=start_1p)) or ((start<=end_1p) and (end>=end_1p)) or ((start<=start_1p) and (end>=end_1p)) or ((start>=start_1p) and (end<=end_1p)):
            return "YES"
        else:
            return "NO"

    elif (chr=="19"):
        if ((start<=start_19q) and (end>=start_19q)) or ((start<=end_19q) and (end>=end_19q)) or ((start<=start_19q) and (end>=end_19q)) or ((start>=start_19q) and (end<=end_19q)):
            return "YES"
        else:
            return "NO"



def get_CNV_file(path, sampleid, ws1, ws2, ws3, ws4):
 

    #Open the relevant CNV file and append to hotspots cnvs tab
    if (os.stat(path+sampleid+"/hotspot_cnvs/"+sampleid+"_1p19q.txt").st_size!=0):
        cnvs=pandas.read_csv(path+sampleid+"/hotspot_cnvs/"+sampleid+"_1p19q.txt", sep="\t")
    else:
        ws3['A1']= 'ERROR- cannot find cnv file'


    #highlight fish probe region using fish_probe function
    cnvs['chromosome']=cnvs['chromosome'].apply(lambda x: str(x))
    cnvs['start']=cnvs['start'].apply(lambda x: str(x))
    cnvs['end']=cnvs['end'].apply(lambda x: str(x))
    cnvs['chr_start'] = cnvs['chromosome'].str.cat(cnvs['start'],sep="-")
    cnvs['chr_start_end']=cnvs['chr_start'].str.cat(cnvs['end'],sep="-")
    cnvs['Fish probe region?']=cnvs['chr_start_end'].apply(lambda x: fish_probe(x))
    cnvs=cnvs.drop(columns=['chr_start', 'chr_start_end', 'n_bins', 'segment_weight'])
    ws3['A9']= ' '
    for row in dataframe_to_rows(cnvs, header=True, index=False):
        ws3.append(row)


    colour= PatternFill("solid", fgColor="00FFFF00")
    row=0
    while (row<cnvs.shape[0]):
        if (cnvs.iloc[row,11]=="YES"):
            ws3["L"+str(row+11)].fill=colour
        row=row+1


    colour= PatternFill("solid", fgColor="00FFFF00")
    row=0
    while (row<cnvs.shape[0]):
        if (cnvs.iloc[row,4]<=-0.4 or cnvs.iloc[row,4]>=0.3):
            ws3["E"+str(row+11)].fill=colour
        row=row+1


    if ((cnvs.shape[0])>10):
        colour= PatternFill("solid", fgColor="00FF0000")
        position= ['A21']
        ws3['A21']="ERROR-CHECK FILE- too many lines!!"
        for cell in position:
            ws3[cell].fill=colour


    ws3['A24']= "Analysis"
    ws3['A26']= "1st checker"
    ws3['A27']= "2nd checker"
    ws3['B25']= "Chr1 Result"
    ws3['C25']= "Chr19 Result"
    ws3['D25']= "Action"
    ws3['E25']= "Name and Date"

    ws3['A30']= "Reporting"
    ws3['A32']= "Result for reporting check 1"
    ws3['A33']= "Result for reporting check 2"
    ws3['B31']= "Results"
    ws3['C31']= "How conclusion reached"
    ws3['D31']= "Additional Comments"
    ws3['E31']= "Name and Date"

    ws3['G25']= "Reference information"
    ws3['G26']= "Reasons to refer to scatter plots:"
    ws3['G27']= "Query loss of 1p and/or 19q relative to gain of control arm i.e. log2 >0.3 for 1q and/or 19p."
    ws3['G28']= "Query loss but threshold not met: Log2 value -0.2 to -0.39 on 1p and 19q arms."
    ws3['G29']= "Query loss outside of FISH probe region: Log2 ratio <-0.4 on 1p +/or 19q but NOT in FISH probe region. NB only if remainder of arm has negative log2 (evidence of loss)"
    ws3['G30']=""
    ws3['G31']= "Reasons to recommend FISH"
    ws3['G32']= "Querying loss of 1p and/or 19q relative to gain of control arm (i.e. log2 >0.3 for 1q and/or 19p). FISH only required if both 1p and 19q show evidence of loss or relative loss (i.e. potential co-deletion)."
    ws3['G33']= "Query loss but threshold not met: 1p and 19q arms appear deleted on scatter plot, and negative log2 values."
    ws3['G34']= "Query loss outside of FISH probe region: Log2 ratio <-0.4 on 1p and/or 19q but NOT in FISH probe region. FISH only required if strong evidence of loss close to the FISH probe region and the other chromosome looks deleted (i.e. potential co-deletion)."
    dv = DataValidation(type="list", formula1='"1p FISH probe deleted, No deletion of 1p FISH probe region, Query 1p relative loss, Deletion of 1p queried- threshold not met in FISH probe region"', allow_blank=True)
    ws3.add_data_validation(dv)
    dv.add('B26')
    dv.add('B27')

    dv = DataValidation(type="list", formula1='"19q FISH probe deleted, No deletion of 19q FISH probe region, Query 19q relative loss, Deletion of 19q queried- threshold not met in FISH probe region"', allow_blank=True)
    ws3.add_data_validation(dv)
    dv.add('C26')
    dv.add('C27')

    dv = DataValidation(type="list", formula1='"Refer to scatter plots, Complete Reporting section"', allow_blank=True)
    ws3.add_data_validation(dv)
    dv.add('D26')
    dv.add('D27')

    ws3['H36']= "p"
    ws3['I36']= "centromere"
    ws3['J36']= "q"

    ws3['G37']= "chr1"
    ws3['H37']= "1-121535433"
    ws3['I37']= "121535434-124535434"
    ws3['J37']= "124535435-249250621"

    ws3['G38']= "chr19"
    ws3['H38']= "1-24681781"
    ws3['I38']= "24681782-27681782"
    ws3['J38']= "27681783-59128983"

    ws3['G40']= "FISH regions"
    ws3['H40']= "Start"
    ws3['I40']= "End"
    ws3['J40']= "size bp"

    ws3['G41']= "1p"
    ws3['H41']= "3397084"
    ws3['I41']= "3766783"
    ws3['J41']= "369699"

    ws3['G42']= "1q"
    ws3['H42']= "178787819"
    ws3['I42']= "179327812"
    ws3['J42']= "539993"

    ws3['G43']= "19p"
    ws3['H43']= "12318204"
    ws3['I43']= "12792515"
    ws3['J43']= "474311"

    ws3['G44']= "19q"
    ws3['H44']= "47924005"
    ws3['I44']= "48374711"
    ws3['J44']= "450706"

    return cnvs, ws1, ws2, ws3, ws4


def get_coverage(path, seqId, sampleid, ws1, ws2, ws3, ws4):

    coverage_NTC=pandas.read_csv(path+ "NTC/"+seqId+ "_NTC_CNVS_DepthOfCoverage", sep ="\t") 
    coverage_NTC=coverage_NTC.drop(columns=['Total_Depth', 'Average_Depth_sample'], axis=1)
    coverage_sample=pandas.read_csv(path+ sampleid +"/"+seqId+"_"+sampleid+ "_CNVS_DepthOfCoverage", sep ="\t") 
    coverage_sample=coverage_sample.drop(columns=['Total_Depth', 'Average_Depth_sample'], axis=1)

    #combine the sample and the NTC depth of coverage file
    coverage=pandas.merge(coverage_sample, coverage_NTC, how="outer", on="Locus")
    coverage['%NTC'] = (coverage['Depth_for_NTC']/coverage['Depth_for_'+sampleid])*100
    for row in dataframe_to_rows(coverage, header=True, index=False):
        ws2.append(row)


    #Highlight %NTC column where values are greater than or equal to 10
    colour= PatternFill("solid", fgColor="00FFFF00")
    row=0
    while (row<coverage.shape[0]):
        if (coverage.iloc[row,3]>=10):
            ws2["E"+str(row+2)].fill=colour
        row=row+1

    ws2['H3']= "Results"
    ws2['I3']= "Comments"
    ws2['G4']= "NTC check 1"
    ws2['G5']= "NTC check 2"
    ws2['G13']= "Total Reads:"
    ws2['G16']= "Total Reads check 1"
    ws2['G17']= "Total Reads check 2"
    ws2['H15']= "Results"
    ws2['I15']= "Comments"
    ws2['I12']= "Note"
    ws2['I13']= "<2,000,000 = Fail"
    ws2['I12'].font = Font(color="00FF0000")
    ws2['I13'].font = Font(color="00FF0000")

    #add a pass/fail dropdown box
    dv = DataValidation(type="list", formula1='"PASS, FAIL"', allow_blank=True)
    ws2.add_data_validation(dv)
    dv.add('H4')
    dv.add('H5')
    dv.add('H16')
    dv.add('H17')

    #get the percentage aligned reads and number of aligned reads
    with open (path+ sampleid+"/"+seqId+ "_"+sampleid+"_AlignmentSummaryMetrics.txt") as file:
        for line in file:
            if line.startswith("CATEGORY"):
                headers=line.split('\t')
            if line.startswith("PAIR"):
                pair_list=line.split('\t')

    alignment_metrics=pandas.DataFrame([pair_list], columns=headers)
    total_reads=alignment_metrics[['TOTAL_READS']]
    total_reads_value=total_reads.iloc[0,0]
    total_reads_value=int(total_reads_value)
    ws2["H13"]=total_reads_value

    #colour the cell red if number of aligned reads is less than the threshold of 20000
    if (total_reads_value<20000):
        colour= PatternFill("solid", fgColor="00FF0000")
        position= ['H13']
        for cell in position:
            ws2[cell].fill=colour

    return coverage, ws1, ws2, ws3, ws4


def format_sheets(path, sampleid, ws1, ws2, ws3, ws4):

    colour= PatternFill("solid", fgColor="0099CC00")
    position= ['A1']
    for cell in position:
        ws1[cell].fill=colour

    border_a=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    position=['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1', 'A8','A9','B8', 'B9', 'A13', 'A14', 'A15', 'A16', 'A17', 'B13', 'B14', 'B15', 'B16', 'B17']
    for cell in position:
        ws1[cell].border=border_a

    colour= PatternFill("solid", fgColor="DCDCDC")
    position= ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1', 'A13', 'A14', 'A15']
    for cell in position:
        ws1[cell].fill=colour

    colour= PatternFill("solid", fgColor="00FFCC00")
    position= ['A8','A9', 'A16', 'A17']
    for cell in position:
        ws1[cell].fill=colour

    ws1.column_dimensions['A'].width=35
    ws1.column_dimensions['B'].width=25
    ws1.column_dimensions['C'].width=25
    ws1.column_dimensions['D'].width=25
    ws1.column_dimensions['E'].width=25
    ws1.column_dimensions['F'].width=25
    ws1.column_dimensions['G'].width=25
    ws1.column_dimensions['H'].width=25
    ws1.column_dimensions['I'].width=25
    ws1.column_dimensions['J'].width=25
    ws1.column_dimensions['K'].width=25
    ws1.column_dimensions['L'].width=25
    ws1.column_dimensions['M'].width=25
    ws1.column_dimensions['N'].width=25
    ws1.column_dimensions['O'].width=25
    ws1.column_dimensions['P'].width=25
    ws1.column_dimensions['Q'].width=25
    ws1.column_dimensions['R'].width=25
    ws1.column_dimensions['S'].width=25
    ws1.column_dimensions['T'].width=25
    ws1.column_dimensions['U'].width=25
    ws1.column_dimensions['V'].width=25

    ws2.column_dimensions['A'].width=23
    ws2.column_dimensions['B'].width=23
    ws2.column_dimensions['C'].width=23
    ws2.column_dimensions['D'].width=23
    ws2.column_dimensions['G'].width=20
    ws2.column_dimensions['H'].width=20
    ws2.column_dimensions['I'].width=40

    border_a=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    position=['G4', 'G5', 'H3', 'H4', 'H5', 'I3', 'I4', 'I5', 'G13', 'G16', 'G17', 'H15', 'H16', 'H17', 'I15', 'I16', 'I17' ]
    for cell in position:
        ws2[cell].border=border_a

    colour= PatternFill("solid", fgColor="DCDCDC")
    position= ['G4','G5','G16','G17']
    for cell in position:
        ws2[cell].fill=colour

    colour= PatternFill("solid", fgColor="00CCFFFF")
    position= ['H3', 'I3', 'G13', 'H15', 'I15']
    for cell in position:
        ws2[cell].fill=colour


    ws3.column_dimensions['G'].width=20

    border_a=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    position=['A4', 'A5', 'B4', 'B5','C4', 'C5', 'D4', 'D5', 'E4', 'E5', 'A7', 'A8', 'B7', 'B8','C7', 'C8', 'D7', 'D8', 'E7', 'E8','F7', 'F8',]
    for cell in position:
        ws3[cell].border=border_a

    border_a=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    position=['A10', 'B10', 'C10', 'D10','E10', 'F10', 'G10', 'H10', 'I10', 'J10','K10','L10',
              'A11', 'B11', 'C11', 'D11','E11', 'F11', 'G11', 'H11', 'I11', 'J11','K11','L11',
              'A12', 'B12', 'C12', 'D12','E12', 'F12', 'G12', 'H12', 'I12', 'J12','K12','L12',
              'A13', 'B13', 'C13', 'D13','E13', 'F13', 'G13', 'H13', 'I13', 'J13','K13','L13',
              'A14', 'B14', 'C14', 'D14','E14', 'F14', 'G14', 'H14', 'I14', 'J14','K14','L14',
              'A15', 'B15', 'C15', 'D15','E15', 'F15', 'G15', 'H15', 'I15', 'J15','K15','L15',
              'A16', 'B16', 'C16', 'D16','E16', 'F16', 'G16', 'H16', 'I16', 'J16','K16','L16',
              'A17', 'B17', 'C17', 'D17','E17', 'F17', 'G17', 'H17', 'I17', 'J17','K17','L17',
              'A18', 'B18', 'C18', 'D18','E18', 'F18', 'G18', 'H18', 'I18', 'J18','K18','L18',
              'A19', 'B19', 'C19', 'D19','E19', 'F19', 'G19', 'H19', 'I19', 'J19','K19','L19',
              'A20', 'B20', 'C20', 'D20','E20', 'F20', 'G20', 'H20', 'I20', 'J20','K20','L20',
              'A24', 'A25', 'B25', 'C25', 'D25', 'E25',
              'A26', 'B26', 'C26', 'D26', 'E26',
              'A27', 'B27', 'C27', 'D27', 'E27',
              'A30', 'A31', 'B31', 'C31', 'D31', 'E31',
              'A32', 'B32', 'C32', 'D32', 'E32',
              'A33', 'B33', 'C33', 'D33', 'E33',
              'G25', 'G26', 'G27', 'G28', 'G29','G30', 'G31', 'G32', 'G33', 'G34',
              'G36', 'H36', 'I36', 'J36',
              'G37', 'H37', 'I37', 'J37',
              'G38', 'H38', 'I38', 'J38',
              'G40', 'H40', 'I40', 'J40',
              'G41', 'H41', 'I41', 'J41',
              'G42', 'H42', 'I42', 'J42',
              'G43', 'H43', 'I43', 'J43',
              'G44', 'H44', 'I44', 'J44']
    for cell in position:
        ws3[cell].border=border_a

    font_bold=Font(bold=True)
    position= ['A3','A4','B4','C4','D4','E4','A7','B7','C7','D7','E7','F7','A10', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'H10', 'I10', 'J10', 'K10', 'L10', 'G25', 'G26', 'G31', 'H36', 'I36','J36','G40','H40', 'I40','J40', 'B25', 'C25', 'D25', 'E25', 'B31', 'C31', 'D31', 'E31', 'A32', 'A33', 'A24', 'A26', 'A27', 'A30']
    for cell in position:
        ws3[cell].font=font_bold

    colour= PatternFill("solid", fgColor="00CCFFFF")
    position= ['A4', 'B4','C4', 'D4', 'E4', 'A7', 'B7','C7', 'D7', 'E7', 'F7', 'B31', 'C31', 'D31', 'E31']
    for cell in position:
        ws3[cell].fill=colour

    colour= PatternFill("solid", fgColor="00FFCC00")
    position= ['A10', 'B10', 'C10', 'D10','E10', 'F10', 'G10', 'H10', 'I10', 'J10','K10','L10', 'B25', 'C25', 'D25', 'E25']
    for cell in position:
        ws3[cell].fill=colour

    
    colour= PatternFill("solid", fgColor="DCDCDC")
    position= ['A32','A33']
    for cell in position:
        ws3[cell].fill=colour

    ws3.merge_cells('G25:K25')
    ws3.merge_cells('G26:K26')
    ws3.merge_cells('G27:K27')
    ws3.merge_cells('G28:K28')
    ws3.merge_cells('G29:K29')
    ws3.merge_cells('G30:K30')
    ws3.merge_cells('G31:K31')
    ws3.merge_cells('G32:K32')
    ws3.merge_cells('G33:K33')
    ws3.merge_cells('G34:K34')
    ws3.merge_cells('G35:K35')

    ws3.column_dimensions['A'].width=25
    ws3.column_dimensions['B'].width=35
    ws3.column_dimensions['C'].width=30
    ws3.column_dimensions['D'].width=30
    ws3.column_dimensions['E'].width=20
    ws3.column_dimensions['F'].width=20
    ws3.column_dimensions['G'].width=15
    ws3.column_dimensions['H'].width=15
    ws3.column_dimensions['I'].width=15
    ws3.column_dimensions['J'].width=15
    ws3.column_dimensions['K'].width=15

    ws3.row_dimensions[26].height=50
    ws3.row_dimensions[27].height=50
    ws3.row_dimensions[28].height=50
    ws3.row_dimensions[29].height=50
    ws3.row_dimensions[31].height=50
    ws3.row_dimensions[32].height=50
    ws3.row_dimensions[33].height=50
    ws3.row_dimensions[34].height=50

    ws3['G25'].alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
    ws3['G26'].alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
    ws3['G27'].alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
    ws3['G28'].alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
    ws3['G29'].alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
    ws3['G31'].alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
    ws3['G32'].alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
    ws3['G33'].alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')
    ws3['G34'].alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')

    ws4.column_dimensions['A'].width=60
    ws4.column_dimensions['B'].width=60

    return ws1, ws2, ws3, ws4


if __name__ == "__main__":

    parser=argparse.ArgumentParser()
    parser.add_argument('--seqId', required=True)
    parser.add_argument('--sampleid', required=True)
    parser.add_argument('--worksheetid', required=True)
    parser.add_argument('--path', required=False)
    args=parser.parse_args()


    seqId=args.seqId
    sampleid=args.sampleid
    worksheetid=args.worksheetid
    path=args.path

    if (path==None):
        path="/data/results/"+seqId+"/RochePanCancer/"



    ws1, ws2, ws3, ws4=populate_cells(seqId, sampleid, worksheetid, ws1, ws2, ws3, ws4)
    cnv_file, ws1, ws2, ws3, ws4=get_CNV_file(path, sampleid, ws1, ws2, ws3, ws4)
    coverage, ws1, ws2, ws3, ws4=get_coverage(path, seqId, sampleid, ws1, ws2, ws3, ws4)
    ws1,ws2,ws3,ws4=format_sheets(path, sampleid, ws1, ws2, ws3, ws4)

    wb.save(path+sampleid+"/hotspot_cnvs/"+sampleid+"_1p19q.xlsx")